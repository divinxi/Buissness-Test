import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "..", "dist", "Automation-ROI-Tracker.xlsx")

INK = "1C1C1E"
ACCENT = "2F5D50"
ACCENT_LIGHT = "E4ECE9"
MUTED = "5A5A60"

CATEGORIES = ["Build Your AI Systems", "Customer Lifecycle Automation",
              "Competitive & Market Intelligence", "Reporting & Ops Rhythm",
              "Delegation & Team Enablement"]
FREQUENCIES = ["Daily", "Weekly", "Monthly", "Quarterly", "One-time"]
FREQ_MULT = {"Daily": 30, "Weekly": 4.33, "Monthly": 1, "Quarterly": 0.33, "One-time": 0}

LAST_ROW = 201  # rows 2..201 usable in Workflow Log


def header_style(ws, row, values, widths=None):
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def build():
    wb = Workbook()

    # --- Instructions sheet ---
    ws0 = wb.active
    ws0.title = "Start Here"
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions["A"].width = 100
    title = ws0.cell(row=2, column=1, value="Automation ROI Tracker — companion to AI Prompt Playbook Vol. 2")
    title.font = Font(bold=True, size=16, color=INK)
    lines = [
        "",
        "How to use this workbook:",
        "1. Every time you turn one of the 25 prompts into a recurring habit (or your own), log it on the 'Workflow Log' tab.",
        "2. Set how often it actually runs (Daily/Weekly/Monthly/Quarterly/One-time) and how many minutes it saves per run.",
        "3. The sheet automatically converts that into monthly minutes saved, since a 5-minute daily habit adds up very differently than a 5-minute quarterly one.",
        "4. The 'ROI Summary' tab totals everything and converts it to a dollar value using your hourly rate.",
        "",
        "Set your hourly rate (or a fair estimate of your time's value) on the ROI Summary tab — everything else calculates itself.",
        "",
        "Unlike a simple time log, this tracker is built for RECURRING systems: a habit you log once but that keeps paying off every week is worth logging accurately, since the frequency is what turns a few minutes into real hours over a year.",
    ]
    for i, line in enumerate(lines, start=3):
        cell = ws0.cell(row=i, column=1, value=line)
        cell.font = Font(bold=line.startswith("How"), size=11, color=INK if not line.startswith(("1.", "2.", "3.", "4.")) else MUTED)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # --- Workflow Log sheet ---
    ws1 = wb.create_sheet("Workflow Log")
    ws1.sheet_view.showGridLines = False
    headers = ["Date Set Up", "Workflow Name", "Category", "Frequency",
               "Time Saved per Run (min)", "Monthly Minutes Saved", "Notes"]
    header_style(ws1, 1, headers, widths=[13, 30, 26, 12, 20, 20, 34])
    ws1.freeze_panes = "A2"

    thin = Side(style="thin", color="D8D6CC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(2, LAST_ROW + 1):
        for c in range(1, 8):
            ws1.cell(row=r, column=c).border = border
        ws1.cell(row=r, column=5).number_format = "0"
        f_cell = ws1.cell(
            row=r, column=6,
            value=f"=IFERROR(E{r}*VLOOKUP(D{r},'ROI Summary'!$G$4:$H$8,2,FALSE),\"\")",
        )
        f_cell.number_format = "0.0"

    dv_cat = DataValidation(type="list", formula1='"' + ",".join(CATEGORIES) + '"', allow_blank=True)
    ws1.add_data_validation(dv_cat)
    dv_cat.add(f"C2:C{LAST_ROW}")

    dv_freq = DataValidation(type="list", formula1='"' + ",".join(FREQUENCIES) + '"', allow_blank=True)
    ws1.add_data_validation(dv_freq)
    dv_freq.add(f"D2:D{LAST_ROW}")

    # --- ROI Summary sheet ---
    ws2 = wb.create_sheet("ROI Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 4
    ws2.column_dimensions["D"].width = 34
    ws2.column_dimensions["E"].width = 18
    ws2.column_dimensions["F"].width = 4
    ws2.column_dimensions["G"].width = 16
    ws2.column_dimensions["H"].width = 18

    t = ws2.cell(row=2, column=1, value="ROI Summary")
    t.font = Font(bold=True, size=16, color=INK)

    ws2.cell(row=4, column=1, value="Your hourly rate ($)").font = Font(bold=True, size=11)
    rate_cell = ws2.cell(row=4, column=2, value=60)
    rate_cell.font = Font(size=11, color=ACCENT, bold=True)
    rate_cell.fill = PatternFill("solid", fgColor=ACCENT_LIGHT)
    rate_cell.number_format = '"$"#,##0'

    ws2.cell(row=6, column=1, value="Total workflows logged").font = Font(bold=True, size=11)
    ws2.cell(row=6, column=2, value=f"=COUNTA('Workflow Log'!B2:B{LAST_ROW})")

    ws2.cell(row=7, column=1, value="Total monthly minutes saved").font = Font(bold=True, size=11)
    ws2.cell(row=7, column=2, value=f"=SUM('Workflow Log'!F2:F{LAST_ROW})")
    ws2.cell(row=7, column=2).number_format = "0"

    ws2.cell(row=8, column=1, value="Total monthly hours saved").font = Font(bold=True, size=11)
    ws2.cell(row=8, column=2, value="=B7/60")
    ws2.cell(row=8, column=2).number_format = "0.0"

    ws2.cell(row=9, column=1, value="Estimated monthly dollar value").font = Font(bold=True, size=12, color=ACCENT)
    monthly_cell = ws2.cell(row=9, column=2, value="=B8*B4")
    monthly_cell.number_format = '"$"#,##0'
    monthly_cell.font = Font(bold=True, size=12, color=ACCENT)

    ws2.cell(row=10, column=1, value="Estimated annual dollar value").font = Font(bold=True, size=12, color=ACCENT)
    annual_cell = ws2.cell(row=10, column=2, value="=B9*12")
    annual_cell.number_format = '"$"#,##0'
    annual_cell.font = Font(bold=True, size=12, color=ACCENT)

    # Per-category breakdown
    header_style(ws2, 3, [None, None, None, "Category", "Monthly Minutes Saved"])
    for i, cat in enumerate(CATEGORIES, start=4):
        ws2.cell(row=i, column=4, value=cat).font = Font(size=10.5)
        formula = f"=SUMIF('Workflow Log'!$C$2:$C${LAST_ROW},D{i},'Workflow Log'!$F$2:$F${LAST_ROW})"
        ws2.cell(row=i, column=5, value=formula)
        ws2.cell(row=i, column=5).number_format = "0.0"

    # Frequency -> monthly multiplier lookup table (used by Workflow Log's VLOOKUP)
    header_style(ws2, 3, [None, None, None, None, None, None, "Frequency", "Monthly Multiplier"])
    for i, freq in enumerate(FREQUENCIES, start=4):
        ws2.cell(row=i, column=7, value=freq).font = Font(size=10.5)
        ws2.cell(row=i, column=8, value=FREQ_MULT[freq]).font = Font(size=10.5)
    note2 = ws2.cell(row=9, column=7, value="Don't edit — Workflow Log looks these up.")
    note2.font = Font(italic=True, size=8.5, color=MUTED)

    note = ws2.cell(row=12, column=1, value="Tip: log frequency honestly — a 'Daily' habit that actually runs a few times a week will overstate your annual total.")
    note.font = Font(italic=True, size=9.5, color=MUTED)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
