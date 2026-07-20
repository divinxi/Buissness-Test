import sys, os
sys.path.insert(0, os.path.dirname(__file__))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "..", "dist", "Invoice-Payment-Tracker.xlsx")

INK = "1C1C1E"
ACCENT = "2F5D50"
ACCENT_LIGHT = "E4ECE9"
MUTED = "5A5A60"
WARN_LIGHT = "F7ECDB"

LOG_ROWS = 300
STATUS_OPTIONS = ["Paid", "Overdue", "Upcoming"]


def header_style(ws, row, values, widths=None):
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def thin_border():
    thin = Side(style="thin", color="D8D6CC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def build():
    wb = Workbook()
    border = thin_border()

    # --- Start Here ---
    ws0 = wb.active
    ws0.title = "Start Here"
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions["A"].width = 104
    title = ws0.cell(row=2, column=1, value="Freelancer Invoice & Late-Payment Toolkit")
    title.font = Font(bold=True, size=16, color=INK)
    lines = [
        "",
        "How to use this workbook:",
        "1. Log every invoice on the 'Invoice Log' tab the day you send it — due date and amount are required for the Status column to work.",
        "2. When a client pays, fill in 'Date Paid' — the Status column automatically flips to Paid and stops counting days overdue.",
        "3. Check the 'Reminder Dashboard' tab any time to see total outstanding, how many invoices are overdue, and which reminder stage each one is due for (see the PDF guide for the actual email scripts).",
        "4. Use the 'Invoice Template' tab as a ready-to-fill invoice for a new client — it calculates the subtotal, tax, and total automatically.",
        "",
        "This workbook is a tracking and reminder-timing tool, not accounting or invoicing software "
        "— it doesn't send invoices or file anything. It exists so you always know, at a glance, "
        "what's outstanding and what to do about it.",
    ]
    for i, line in enumerate(lines, start=3):
        cell = ws0.cell(row=i, column=1, value=line)
        is_step = line[:2] in ("1.", "2.", "3.", "4.")
        cell.font = Font(bold=line.startswith("How"), size=11, color=INK if not is_step else MUTED)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # --- Invoice Log ---
    ws1 = wb.create_sheet("Invoice Log")
    ws1.sheet_view.showGridLines = False
    header_style(ws1, 1, ["Date Sent", "Client", "Invoice #", "Description", "Amount",
                           "Due Date", "Date Paid", "Status", "Days Overdue"],
                 widths=[13, 22, 12, 30, 13, 13, 13, 12, 13])
    ws1.freeze_panes = "A2"
    dv_status = DataValidation(type="list", formula1='"Paid,Overdue,Upcoming"', allow_blank=True)
    ws1.add_data_validation(dv_status)
    for r in range(2, LOG_ROWS + 2):
        for c in range(1, 10):
            ws1.cell(row=r, column=c).border = border
        ws1.cell(row=r, column=5).number_format = '"$"#,##0.00'
        status_formula = f'=IF(G{r}<>"","Paid",IF(AND(F{r}<>"",F{r}<TODAY()),"Overdue",IF(F{r}<>"","Upcoming","")))'
        ws1.cell(row=r, column=8, value=status_formula)
        overdue_formula = f'=IF(AND(G{r}="",F{r}<>"",F{r}<TODAY()),TODAY()-F{r},"")'
        ws1.cell(row=r, column=9, value=overdue_formula)
    note = ws1.cell(row=LOG_ROWS + 3, column=1,
                     value="Status and Days Overdue calculate automatically from Due Date and Date Paid — don't type into those two columns.")
    note.font = Font(italic=True, size=9.5, color=MUTED)

    # --- Reminder Dashboard ---
    ws2 = wb.create_sheet("Reminder Dashboard")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 4
    ws2.column_dimensions["D"].width = 46

    dt = ws2.cell(row=2, column=1, value="Reminder Dashboard")
    dt.font = Font(bold=True, size=16, color=INK)

    rng_amt = f"'Invoice Log'!$E$2:$E${LOG_ROWS + 1}"
    rng_paid = f"'Invoice Log'!$G$2:$G${LOG_ROWS + 1}"
    rng_status = f"'Invoice Log'!$H$2:$H${LOG_ROWS + 1}"
    rng_overdue = f"'Invoice Log'!$I$2:$I${LOG_ROWS + 1}"

    ws2.cell(row=4, column=1, value="Total invoiced (all time)").font = Font(bold=True, size=11)
    v = ws2.cell(row=4, column=2, value=f"=SUM({rng_amt})")
    v.number_format = '"$"#,##0.00'

    ws2.cell(row=5, column=1, value="Total outstanding (unpaid)").font = Font(bold=True, size=12, color=ACCENT)
    v = ws2.cell(row=5, column=2, value=f'=SUMIF({rng_paid},"",{rng_amt})')
    v.number_format = '"$"#,##0.00'
    v.font = Font(bold=True, size=12, color=ACCENT)

    ws2.cell(row=7, column=1, value="Invoices overdue").font = Font(bold=True, size=11)
    v = ws2.cell(row=7, column=2, value=f'=COUNTIF({rng_status},"Overdue")')

    ws2.cell(row=8, column=1, value="Invoices upcoming (not yet due)").font = Font(size=11)
    v = ws2.cell(row=8, column=2, value=f'=COUNTIF({rng_status},"Upcoming")')

    ws2.cell(row=9, column=1, value="Invoices paid").font = Font(size=11)
    v = ws2.cell(row=9, column=2, value=f'=COUNTIF({rng_status},"Paid")')

    ws2.cell(row=11, column=1, value="Overdue 1-6 days (send: Due date / 7-day reminder next)").font = Font(size=10.5)
    v = ws2.cell(row=11, column=2, value=f'=COUNTIFS({rng_overdue},">=1",{rng_overdue},"<7")')

    ws2.cell(row=12, column=1, value="Overdue 7-13 days (send: 7-day firmer follow-up)").font = Font(size=10.5)
    v = ws2.cell(row=12, column=2, value=f'=COUNTIFS({rng_overdue},">=7",{rng_overdue},"<14")')

    ws2.cell(row=13, column=1, value="Overdue 14+ days (send: final notice / consider escalation)").font = Font(bold=True, size=10.5)
    v = ws2.cell(row=13, column=2, value=f'=COUNTIFS({rng_overdue},">=14")')
    v.font = Font(bold=True, size=10.5, color=ACCENT)

    ws2.cell(row=4, column=4, value="How to use this tab").font = Font(bold=True, size=11)
    tip_lines = [
        "The counts on the left tell you how many invoices need which reminder stage today.",
        "Filter the Invoice Log's Days Overdue column to find exactly which invoices they are.",
        "Reminder email scripts for each stage are in the PDF guide.",
        "Check this tab on a fixed schedule (weekly is enough) — don't wait until you happen to remember an invoice is late.",
    ]
    for i, line in enumerate(tip_lines, start=5):
        c = ws2.cell(row=i, column=4, value=line)
        c.font = Font(size=10, color=MUTED)
        c.alignment = Alignment(wrap_text=True)

    # --- Invoice Template ---
    ws3 = wb.create_sheet("Invoice Template")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 14

    ws3.cell(row=2, column=1, value="INVOICE").font = Font(bold=True, size=20, color=ACCENT)

    labels = [
        (4, "From (your business name)"), (5, "Your address / email"),
        (7, "Bill To (client / company)"), (8, "Client address / email"),
        (10, "Invoice #"), (11, "Date Issued"), (12, "Due Date"), (13, "Payment Terms"),
    ]
    for row, label in labels:
        ws3.cell(row=row, column=1, value=label).font = Font(size=10.5, color=MUTED)
        ws3.cell(row=row, column=2).font = Font(size=11, color=INK)
        ws3.cell(row=row, column=2).border = border

    header_style(ws3, 16, ["Description", "Qty", "Rate", "Amount"])
    for r in range(17, 27):
        for c in range(1, 5):
            ws3.cell(row=r, column=c).border = border
        ws3.cell(row=r, column=3).number_format = '"$"#,##0.00'
        ws3.cell(row=r, column=4).number_format = '"$"#,##0.00'
        ws3.cell(row=r, column=4, value=f'=IF(OR(B{r}="",C{r}=""),"",B{r}*C{r})')

    ws3.cell(row=28, column=3, value="Subtotal").font = Font(bold=True, size=10.5)
    sub_cell = ws3.cell(row=28, column=4, value="=SUM(D17:D27)")
    sub_cell.number_format = '"$"#,##0.00'

    ws3.cell(row=29, column=3, value="Tax %").font = Font(size=10.5)
    tax_pct = ws3.cell(row=29, column=4, value=0)
    tax_pct.number_format = '0.0"%"'

    ws3.cell(row=30, column=3, value="Total Due").font = Font(bold=True, size=12, color=ACCENT)
    total_cell = ws3.cell(row=30, column=4, value="=D28*(1+D29/100)")
    total_cell.number_format = '"$"#,##0.00'
    total_cell.font = Font(bold=True, size=12, color=ACCENT)

    ws3.cell(row=33, column=1, value="Payment methods accepted").font = Font(bold=True, size=10.5)
    ws3.cell(row=34, column=1, value="").border = border
    ws3.cell(row=34, column=1).font = Font(size=10.5)

    ws3.cell(row=36, column=1, value="Late fee terms (state this if you use one — see PDF guide)").font = Font(bold=True, size=10.5)
    ws3.cell(row=37, column=1, value="").border = border

    tip = ws3.cell(row=40, column=1,
                    value="Fill in the highlighted cells and either export this tab (File > Export/Print to PDF) or copy the layout into your invoicing tool of choice.")
    tip.font = Font(italic=True, size=9.3, color=MUTED)
    tip.alignment = Alignment(wrap_text=True)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
