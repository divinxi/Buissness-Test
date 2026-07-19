import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "..", "dist", "AI-Ops-Tracker.xlsx")

INK = "1C1C1E"
ACCENT = "2F5D50"
ACCENT_LIGHT = "E4ECE9"
MUTED = "5A5A60"

CATEGORIES = ["Marketing & Content", "Customer Service & Support", "Finance & Operations",
              "Hiring & HR", "Sales & Outreach"]

def header_style(ws, row, values, widths=None):
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

def build():
    wb = Workbook()

    # --- Instructions sheet ---
    ws0 = wb.active
    ws0.title = "Start Here"
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions["A"].width = 100
    title = ws0.cell(row=2, column=1, value="AI Ops Tracker — companion to the Small Business AI Prompt Playbook")
    title.font = Font(bold=True, size=16, color=INK)
    lines = [
        "",
        "How to use this workbook:",
        "1. Every time you use one of the 25 prompts from the playbook (or your own), log it on the 'Prompt Log' tab.",
        "2. Estimate the time it saved you vs. doing the task manually — be honest, rough is fine.",
        "3. The 'ROI Summary' tab automatically totals your time saved and converts it to a dollar value using your hourly rate.",
        "4. Revisit monthly — it's the fastest way to see which categories are actually paying off, so you know where to lean in.",
        "",
        "Set your hourly rate (or a fair estimate of your time's value) on the ROI Summary tab — everything else calculates itself.",
    ]
    for i, line in enumerate(lines, start=3):
        cell = ws0.cell(row=i, column=1, value=line)
        cell.font = Font(bold=line.startswith("How"), size=11, color=INK if not line.startswith(("1.", "2.", "3.", "4.")) else MUTED)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # --- Prompt Log sheet ---
    ws1 = wb.create_sheet("Prompt Log")
    ws1.sheet_view.showGridLines = False
    headers = ["Date", "Category", "Prompt Used", "Time Saved (minutes)", "Notes / Result"]
    header_style(ws1, 1, headers, widths=[13, 26, 32, 20, 40])
    ws1.freeze_panes = "A2"

    thin = Side(style="thin", color="D8D6CC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(2, 202):
        for c in range(1, 6):
            ws1.cell(row=r, column=c).border = border
        ws1.cell(row=r, column=4).number_format = "0"

    dv_cat = DataValidation(type="list", formula1='"' + ",".join(CATEGORIES) + '"', allow_blank=True)
    ws1.add_data_validation(dv_cat)
    dv_cat.add(f"B2:B201")

    # --- ROI Summary sheet ---
    ws2 = wb.create_sheet("ROI Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 4
    ws2.column_dimensions["D"].width = 34
    ws2.column_dimensions["E"].width = 18

    t = ws2.cell(row=2, column=1, value="ROI Summary")
    t.font = Font(bold=True, size=16, color=INK)

    ws2.cell(row=4, column=1, value="Your hourly rate ($)").font = Font(bold=True, size=11)
    rate_cell = ws2.cell(row=4, column=2, value=50)
    rate_cell.font = Font(size=11, color=ACCENT, bold=True)
    rate_cell.fill = PatternFill("solid", fgColor=ACCENT_LIGHT)
    rate_cell.number_format = '"$"#,##0'

    ws2.cell(row=6, column=1, value="Total prompts logged").font = Font(bold=True, size=11)
    ws2.cell(row=6, column=2, value='=COUNTA(\'Prompt Log\'!A2:A201)')

    ws2.cell(row=7, column=1, value="Total minutes saved").font = Font(bold=True, size=11)
    ws2.cell(row=7, column=2, value="=SUM('Prompt Log'!D2:D201)")

    ws2.cell(row=8, column=1, value="Total hours saved").font = Font(bold=True, size=11)
    ws2.cell(row=8, column=2, value="=B7/60")
    ws2.cell(row=8, column=2).number_format = "0.0"

    ws2.cell(row=9, column=1, value="Estimated dollar value saved").font = Font(bold=True, size=12, color=ACCENT)
    dollar_cell = ws2.cell(row=9, column=2, value="=B8*B4")
    dollar_cell.number_format = '"$"#,##0'
    dollar_cell.font = Font(bold=True, size=12, color=ACCENT)

    # Per-category breakdown
    ws2.cell(row=4, column=4, value="By category").font = Font(bold=True, size=11)
    header_style(ws2, 3, [None, None, None, "Category", "Minutes Saved"])
    for i, cat in enumerate(CATEGORIES, start=4):
        ws2.cell(row=i, column=4, value=cat).font = Font(size=10.5)
        formula = f"=SUMIF('Prompt Log'!$B$2:$B$201,D{i},'Prompt Log'!$D$2:$D$201)"
        ws2.cell(row=i, column=5, value=formula)

    note = ws2.cell(row=12, column=1, value="Tip: log time saved conservatively — this number is most useful when you trust it.")
    note.font = Font(italic=True, size=9.5, color=MUTED)

    wb.save(OUT)
    print(f"Wrote {OUT}")

if __name__ == "__main__":
    build()
