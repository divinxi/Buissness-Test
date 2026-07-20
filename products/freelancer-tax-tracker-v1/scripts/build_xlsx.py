import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from content import TAX_YEAR, QUARTERLY_DEADLINES, SE_TAX_FACTS, SCHEDULE_C_CATEGORIES

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "..", "dist", "Freelancer-Tax-Tracker.xlsx")

INK = "1C1C1E"
ACCENT = "2F5D50"
ACCENT_LIGHT = "E4ECE9"
MUTED = "5A5A60"
WARN_LIGHT = "F7ECDB"

CATEGORY_NAMES = [c["name"] for c in SCHEDULE_C_CATEGORIES] + ["Home Office", "Other"]
LOG_ROWS = 300  # data rows provisioned in Income/Expense logs


def header_style(ws, row, values, widths=None):
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def thin_border():
    thin = Side(style="thin", color="D8D6CC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def build():
    wb = Workbook()
    border = thin_border()

    # --- Start Here ---
    ws0 = wb.active
    ws0.title = "Start Here"
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions["A"].width = 104
    title = ws0.cell(row=2, column=1, value=f"Freelancer Quarterly Tax & Expense Tracker — {TAX_YEAR} Edition")
    title.font = Font(bold=True, size=16, color=INK)
    lines = [
        "",
        "How to use this workbook:",
        "1. Log every client payment on the 'Income Log' tab as it arrives.",
        "2. Log every business expense on the 'Expense Log' tab — pick a category from the dropdown; it maps to real IRS Schedule C lines (see the PDF guide).",
        "3. Meals default to 50% deductible — the sheet applies that automatically; every other category defaults to 100% but you can override the % on any row.",
        "4. Open the 'Quarterly Tax Estimator' tab and type your projected full-year net profit — it calculates your exact self-employment tax using the real 2026 formula.",
        "5. Check the 'Dashboard' tab any time for a live snapshot: income, expenses, net profit, and how many days until your next quarterly payment is due.",
        "",
        "This workbook is a bookkeeping and estimating tool, not tax software — it does not file "
        "anything or replace a preparer. It exists so your numbers are already organized when tax time comes.",
    ]
    for i, line in enumerate(lines, start=3):
        cell = ws0.cell(row=i, column=1, value=line)
        is_step = line[:2] in ("1.", "2.", "3.", "4.", "5.")
        cell.font = Font(bold=line.startswith("How"), size=11, color=INK if not is_step else MUTED)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # --- Income Log ---
    ws1 = wb.create_sheet("Income Log")
    ws1.sheet_view.showGridLines = False
    header_style(ws1, 1, ["Date", "Client", "Invoice #", "Description", "Amount", "Date Paid"],
                 widths=[13, 24, 14, 32, 14, 14])
    ws1.freeze_panes = "A2"
    for r in range(2, LOG_ROWS + 2):
        for c in range(1, 7):
            ws1.cell(row=r, column=c).border = border
        ws1.cell(row=r, column=5).number_format = '"$"#,##0.00'

    # --- Expense Log ---
    ws2 = wb.create_sheet("Expense Log")
    ws2.sheet_view.showGridLines = False
    header_style(ws2, 1, ["Date", "Category", "Vendor", "Amount", "Deductible %", "Deductible Amount", "Notes"],
                 widths=[13, 26, 22, 13, 13, 17, 34])
    ws2.freeze_panes = "A2"
    dv_cat = DataValidation(type="list", formula1='"' + ",".join(CATEGORY_NAMES) + '"', allow_blank=True)
    ws2.add_data_validation(dv_cat)
    dv_cat.add(f"B2:B{LOG_ROWS + 1}")
    for r in range(2, LOG_ROWS + 2):
        for c in range(1, 8):
            ws2.cell(row=r, column=c).border = border
        ws2.cell(row=r, column=4).number_format = '"$"#,##0.00'
        pct_cell = ws2.cell(row=r, column=5, value=f'=IF(B{r}="Meals",50,IF(B{r}="","",100))')
        pct_cell.number_format = '0"%"'
        ws2.cell(row=r, column=6, value=f'=IF(OR(D{r}="",B{r}=""),"",D{r}*(E{r}/100))').number_format = '"$"#,##0.00'
    note = ws2.cell(row=LOG_ROWS + 3, column=1,
                     value="Deductible % auto-fills to 50% for Meals, 100% otherwise — overwrite any cell if your real business-use share is different (e.g. a partly-personal vehicle or phone line).")
    note.font = Font(italic=True, size=9.5, color=MUTED)

    # --- Quarterly Tax Estimator ---
    ws3 = wb.create_sheet("Quarterly Tax Estimator")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 4
    ws3.column_dimensions["D"].width = 46

    t = ws3.cell(row=2, column=1, value="Quarterly Tax Estimator")
    t.font = Font(bold=True, size=16, color=INK)

    ws3.cell(row=4, column=1, value="Net profit YTD (from logs)").font = Font(bold=True, size=11)
    ytd_cell = ws3.cell(row=4, column=2,
                         value="=SUM('Income Log'!E2:E{r}) - SUM('Expense Log'!F2:F{r})".format(r=LOG_ROWS + 1))
    ytd_cell.number_format = '"$"#,##0.00'

    ws3.cell(row=5, column=1, value="Projected full-year net profit (edit this)").font = Font(bold=True, size=11, color=ACCENT)
    proj_cell = ws3.cell(row=5, column=2, value=60000)
    proj_cell.font = Font(size=11, color=ACCENT, bold=True)
    proj_cell.fill = PatternFill("solid", fgColor=ACCENT_LIGHT)
    proj_cell.number_format = '"$"#,##0.00'

    wage_base = SE_TAX_FACTS["wage_base_2026"]
    ss_rate = SE_TAX_FACTS["social_security_rate"]
    med_rate = SE_TAX_FACTS["medicare_rate"]
    net_factor = SE_TAX_FACTS["net_earnings_factor"]

    rows = [
        (7, "Net earnings from self-employment", f"=B5*{net_factor}", '"$"#,##0.00'),
        (8, f"Social Security tax ({ss_rate*100:.1f}%, capped at ${wage_base:,} wage base)",
         f"=MIN(B7,{wage_base})*{ss_rate}", '"$"#,##0.00'),
        (9, f"Medicare tax ({med_rate*100:.1f}%, no cap)", f"=B7*{med_rate}", '"$"#,##0.00'),
        (10, "Total self-employment tax", "=B8+B9", '"$"#,##0.00'),
        (11, "Deductible half of SE tax (for income tax return)", "=B10/2", '"$"#,##0.00'),
    ]
    for r, label, formula, fmt in rows:
        lbl = ws3.cell(row=r, column=1, value=label)
        lbl.font = Font(bold=(r == 10), size=11, color=ACCENT if r == 10 else INK)
        val = ws3.cell(row=r, column=2, value=formula)
        val.number_format = fmt
        val.font = Font(bold=(r == 10), size=11, color=ACCENT if r == 10 else INK)

    ws3.cell(row=13, column=1, value="Suggested total tax reserve (25%-30% rule of thumb)").font = Font(bold=True, size=11)
    ws3.cell(row=14, column=1, value="  Low estimate (25% of projected net profit)").font = Font(size=10.5)
    low_cell = ws3.cell(row=14, column=2, value="=B5*0.25")
    low_cell.number_format = '"$"#,##0.00'
    ws3.cell(row=15, column=1, value="  High estimate (30% of projected net profit)").font = Font(size=10.5)
    high_cell = ws3.cell(row=15, column=2, value="=B5*0.30")
    high_cell.number_format = '"$"#,##0.00'

    ws3.cell(row=17, column=1, value="Suggested payment THIS quarter (reserve ÷ 4)").font = Font(bold=True, size=12, color=ACCENT)
    q_low = ws3.cell(row=17, column=2, value="=B14/4")
    q_low.number_format = '"$"#,##0.00'
    q_low.font = Font(bold=True, size=12, color=ACCENT)
    ws3.cell(row=18, column=1, value="  to").font = Font(size=10.5, color=MUTED)
    q_high = ws3.cell(row=18, column=2, value="=B15/4")
    q_high.number_format = '"$"#,##0.00'
    q_high.font = Font(size=11, color=ACCENT)

    # Deadlines block
    ws3.cell(row=4, column=4, value=f"{TAX_YEAR} Payment Deadlines").font = Font(bold=True, size=11)
    header_style(ws3, 5, [None, None, None, "Quarter", "Due Date"])
    ws3.column_dimensions["E"].width = 16
    for i, d in enumerate(QUARTERLY_DEADLINES, start=6):
        ws3.cell(row=i, column=4, value=d["quarter"]).font = Font(size=10.5)
        due_cell = ws3.cell(row=i, column=5, value=d["due"])
        due_cell.font = Font(size=10.5)

    days_lbl = ws3.cell(row=11, column=4, value="Days until next deadline")
    days_lbl.font = Font(bold=True, size=10.5)
    days_formula = (
        '=MIN(IF(DATE(2026,4,15)>=TODAY(),DATE(2026,4,15)-TODAY(),9999),'
        'IF(DATE(2026,6,15)>=TODAY(),DATE(2026,6,15)-TODAY(),9999),'
        'IF(DATE(2026,9,15)>=TODAY(),DATE(2026,9,15)-TODAY(),9999),'
        'IF(DATE(2027,1,15)>=TODAY(),DATE(2027,1,15)-TODAY(),9999))'
    )
    days_cell = ws3.cell(row=11, column=5, value=days_formula)
    days_cell.font = Font(bold=True, size=11, color=ACCENT)

    caveat = ws3.cell(row=20, column=1,
                       value="This tab calculates self-employment tax exactly (it's fixed math). The 25%-30% total "
                             "reserve is a rule-of-thumb range that also covers income tax — your real income-tax "
                             "rate depends on filing status, deductions, and state. Use the IRS Form 1040-ES worksheet "
                             "or a tax preparer for a personalized income-tax figure.")
    caveat.font = Font(italic=True, size=9.3, color=MUTED)
    caveat.alignment = Alignment(wrap_text=True)
    ws3.row_dimensions[20].height = 40

    # --- Dashboard ---
    ws4 = wb.create_sheet("Dashboard")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 34
    ws4.column_dimensions["B"].width = 18
    ws4.column_dimensions["C"].width = 4
    ws4.column_dimensions["D"].width = 30
    ws4.column_dimensions["E"].width = 16

    dt = ws4.cell(row=2, column=1, value="Dashboard")
    dt.font = Font(bold=True, size=16, color=INK)

    ws4.cell(row=4, column=1, value="Total income YTD").font = Font(bold=True, size=11)
    inc_cell = ws4.cell(row=4, column=2, value=f"=SUM('Income Log'!E2:E{LOG_ROWS + 1})")
    inc_cell.number_format = '"$"#,##0.00'

    ws4.cell(row=5, column=1, value="Total deductible expenses YTD").font = Font(bold=True, size=11)
    exp_cell = ws4.cell(row=5, column=2, value=f"=SUM('Expense Log'!F2:F{LOG_ROWS + 1})")
    exp_cell.number_format = '"$"#,##0.00'

    ws4.cell(row=6, column=1, value="Net profit YTD").font = Font(bold=True, size=12, color=ACCENT)
    net_cell = ws4.cell(row=6, column=2, value="=B4-B5")
    net_cell.number_format = '"$"#,##0.00'
    net_cell.font = Font(bold=True, size=12, color=ACCENT)

    ws4.cell(row=8, column=1, value="Suggested quarterly payment (from Estimator tab)").font = Font(bold=True, size=11)
    qp_cell = ws4.cell(row=8, column=2, value="='Quarterly Tax Estimator'!B17")
    qp_cell.number_format = '"$"#,##0.00'

    ws4.cell(row=9, column=1, value="Days until next deadline").font = Font(bold=True, size=11)
    dd_cell = ws4.cell(row=9, column=2, value="='Quarterly Tax Estimator'!E11")
    dd_cell.font = Font(bold=True, size=11, color=ACCENT)

    # Per-category breakdown
    ws4.cell(row=4, column=4, value="Expenses by Category (YTD)").font = Font(bold=True, size=11)
    header_style(ws4, 3, [None, None, None, "Category", "Amount"])
    for i, cat in enumerate(CATEGORY_NAMES, start=4):
        ws4.cell(row=i, column=4, value=cat).font = Font(size=10)
        formula = f"=SUMIF('Expense Log'!$B$2:$B${LOG_ROWS + 1},D{i},'Expense Log'!$F$2:$F${LOG_ROWS + 1})"
        cell = ws4.cell(row=i, column=5, value=formula)
        cell.number_format = '"$"#,##0.00'

    tip = ws4.cell(row=4 + len(CATEGORY_NAMES) + 2, column=1,
                    value="Update your logs weekly — this dashboard is only as current as your last entry.")
    tip.font = Font(italic=True, size=9.5, color=MUTED)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
